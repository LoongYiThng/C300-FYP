import sys
from docx import Document

if sys.argv[1]=="docx":
    def suppress_data(text):
        suppressed_text = ''
        words = text.split()
        for word in words:
            if is_name(word):  # Suppress names
                suppressed_text += ''  # Suppress the name by not appending it
            elif is_email(word):  # Suppress email addresses
                suppressed_text += ''  # Suppress the email address by not appending it
            elif is_number(word):  # Suppress numbers (including phone numbers)
                suppressed_text += ''  # Suppress the number by not appending it
            else:
                suppressed_text += word + ' '  # Append non-sensitive words
        return suppressed_text.strip()

    def is_name(word):
        # Check if the word is likely to be a name
        return word[0].isalpha() and word[0].isupper()

    def is_email(word):
        # Check if the word is likely to be an email address
        return '@' in word and '.' in word

    def is_number(word):
        # Check if the word is likely to be a number (including phone numbers)
        return word.isdigit() or is_phone_number(word)

    def is_phone_number(word):
        # Check if the word is likely to be a phone number
        return any(char.isdigit() for char in word)

    # Open the Word document
    doc = Document(sys.argv[2])

    # Process each paragraph in the document
    for paragraph in doc.paragraphs:
        original_text = paragraph.text
        suppressed_text = suppress_data(original_text)
        paragraph.text = suppressed_text

    # Save the modified document
    doc.save(sys.argv[2])
    print("Done! Check save folder.")

elif sys.argv[1]=="xlsx":
    import os
    import random
    from openpyxl import load_workbook

    def suppress_names(sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    words = cell.value.split()
                    suppressed_words = ['X' * len(word) if word.isalpha() else word for word in words]
                    suppressed_value = ' '.join(suppressed_words)
                    cell.value = suppressed_value


    def suppress_numbers(sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.value = 'N/A'


    def suppress_phone_numbers(sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(char.isdigit() for char in cell.value):
                    suppressed_value = ''.join('X' if char.isdigit() else char for char in cell.value)
                    cell.value = suppressed_value


    def suppress_email_addresses(sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and '@' in cell.value:
                    cell.value = 'X' * len(cell.value)


    def suppress_attributes_xlsx_file(filename):
        workbook = load_workbook(filename)
        for sheet in workbook:
            suppress_names(sheet)
            suppress_numbers(sheet)
            suppress_phone_numbers(sheet)
            suppress_email_addresses(sheet)
        save_filename=rename_file_extension(filename, ".tmp")
        workbook.save(save_filename)
        print("save: "+save_filename)

    def rename_file_extension(old_filename, new_extension):
        name, old_extension=os.path.splitext(old_filename)
        new_filename=name+new_extension
        os.rename(old_filename, new_filename)

        return new_filename

    # Usage example
    filename = sys.argv[2]

    old_file=filename
    new_extension=".xlsx"
    new_filename=rename_file_extension(old_file, new_extension)
    print("old: "+old_file)
    print("new: "+new_filename)
    

    suppress_attributes_xlsx_file(new_filename)
    print("Done")