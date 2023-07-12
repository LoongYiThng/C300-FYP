import sys
from docx import Document
from openpyxl import load_workbook

if sys.argv[1]=="document":
    def mask_data(text):
        masked_text = ''
        words = text.split()
        for word in words:
            if is_name(word):  # Mask names
                masked_text += mask_name(word) + ' '
            elif is_email(word):  # Mask email addresses
                masked_text += mask_email(word) + ' '
            elif is_number(word):  # Mask numbers (including phone numbers)
                masked_text += mask_number(word) + ' '
            else:
                masked_text += word + ' '
        return masked_text.strip()

    def is_name(word):
        # Check if the word is likely to be a name
        return word[0].isalpha() and word[0].isupper()

    def is_email(word):
        # Check if the word is likely to be an email address
        return '@' in word and '.' in word

    def is_number(word):
        # Check if the word is likely to be a number (including phone numbers)
        return word.isdigit() or is_phone_number(word)

    def is_phone_number(word):
        # Check if the word is likely to be a phone number
        return any(char.isdigit() for char in word)

    def mask_name(name):
        # Mask a name by replacing characters with 'X'
        masked_name = ''
        for char in name:
            if char.isalnum():
                masked_name += 'X'
            else:
                masked_name += char
        return masked_name

    def mask_email(email):
        # Mask an email address by replacing the entire email address with 'X'
        return 'X' * len(email)

    def mask_number(number):
        # Mask a number (including phone numbers) by replacing digits with 'X'
        return 'X' * len(number)

    # Open the Word document
    doc = Document('test.docx')

    # Process each paragraph in the document
    for paragraph in doc.paragraphs:
        original_text = paragraph.text
        masked_text = mask_data(original_text)
        paragraph.text = masked_text

    # Save the modified document
    doc.save('masked.docx')
    print("Done! Check save folder")

elif sys.argv[1]=="excel":
    def mask_string(string):
        return '*' * len(string)


    def mask_names(sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    words = cell.value.split()
                    masked_words = [mask_string(word) for word in words]
                    masked_value = ' '.join(masked_words)
                    cell.value = masked_value


    def mask_numbers(sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.value = mask_string(str(cell.value))


    def mask_phone_numbers(sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(char.isdigit() for char in cell.value):
                    masked_value = ''.join(mask_string(char) if char.isdigit() else char for char in cell.value)
                    cell.value = masked_value


    def mask_email_addresses(sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and '@' in cell.value:
                    parts = cell.value.split('@')
                    masked_username = mask_string(parts[0])
                    masked_value = masked_username + '@' + parts[1]
                    cell.value = masked_value


    def mask_xlsx_file(filename):
        workbook = load_workbook(filename)
        for sheet in workbook:
            mask_names(sheet)
            mask_numbers(sheet)
            mask_phone_numbers(sheet)
            mask_email_addresses(sheet)
        workbook.save(f"masked_{filename}")


    # Usage example
    filename = "test.xlsx"
    mask_xlsx_file(filename)
    print("Done")