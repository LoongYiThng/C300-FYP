import sys
from faker import Faker
from openpyxl import load_workbook

if sys.argv[1]=="document":
    from docx import Document

    faker = Faker()

    # Example original document
    document = Document('test.docx')

    # Function to sanitize text
    def sanitize_text(text):
        # Split the text into words
        words = text.split()
        # Generate a fake word for each original word
        fake_words = [faker.word() for _ in words]
        # Combine the fake words into a string
        fake_text = ' '.join(fake_words)
        return fake_text

    # Sanitize the document
    for paragraph in document.paragraphs:
        # Sanitize the text in the paragraph
        sanitized_text = sanitize_text(paragraph.text)
        # Replace the original text with the sanitized text
        paragraph.text = sanitized_text

    # Save the sanitized document
    document.save('synthetic.docx')
    print("Done, Check save folder")

elif sys.argv[1]=="excel":
    faker = Faker()

    # Example original XLSX file
    filename = 'test.xlsx'

    # Function to sanitize data
    def sanitize_data(data):
        # Generate fake data based on the original data type
        if isinstance(data, str):
            return faker.word()
        elif isinstance(data, int):
            return faker.random_number()
        elif isinstance(data, float):
            return faker.random_number(digits=2)
        else:
            return data

    # Load the workbook
    wb = load_workbook(filename)

    # Iterate over each sheet in the workbook
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Iterate over each cell in the sheet
        for row in ws.iter_rows():
            for cell in row:
                # Sanitize the cell value
                cell.value = sanitize_data(cell.value)

    # Save the sanitized workbook
    sanitized_filename = 'sanitized_' + filename
    wb.save(sanitized_filename)
    print("Done")