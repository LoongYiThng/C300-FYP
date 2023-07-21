import sys
from docx import Document
import random
import string
from openpyxl import load_workbook

if sys.argv[1]=="docx":
    def perturb_data(text):
        perturbed_text = ''
        words = text.split()
        for word in words:
            if is_name(word):  # Perturb names
                perturbed_text += perturb_name(word) + ' '
            elif is_email(word):  # Perturb email addresses
                perturbed_text += perturb_email(word) + ' '
            elif is_number(word):  # Perturb numbers (including phone numbers)
                perturbed_text += perturb_number(word) + ' '
            else:
                perturbed_text += word + ' '
        return perturbed_text.strip()

    def is_name(word):
        # Check if the word is likely to be a name
        return word[0].isalpha() and word[0].isupper()

    def is_email(word):
        # Check if the word is likely to be an email address
        return '@' in word and '.' in word and word.index('@') < word.rindex('.')

    def is_number(word):
        # Check if the word is likely to be a number (including phone numbers)
        return word.isdigit() or is_phone_number(word)

    def is_phone_number(word):
        # Check if the word is likely to be a phone number
        return any(char.isdigit() for char in word)

    def perturb_name(name):
        # Perturb a name by randomly replacing characters
        perturbed_name = ''
        for char in name:
            if char.isalnum():
                perturbed_name += random.choice(string.ascii_letters)
            else:
                perturbed_name += char
        return perturbed_name

    def perturb_email(email):
        # Perturb an email address by randomly changing characters before the '@' symbol
        username, domain = email.split('@')
        perturbed_username = perturb_name(username)
        return perturbed_username + '@' + domain

    def perturb_number(number):
        # Perturb a number (including phone numbers) by randomly changing digits
        perturbed_number = ''
        for digit in number:
            if digit.isdigit():
                perturbed_number += random.choice(string.digits)
            else:
                perturbed_number += digit
        return perturbed_number

    # Open the Word document
    doc = Document(sys.argv[2])

    # Process each paragraph in the document
    for paragraph in doc.paragraphs:
        original_text = paragraph.text
        perturbed_text = perturb_data(original_text)
        paragraph.text = perturbed_text

    # Save the modified document
    doc.save(sys.argv[2])
    print("Done! Check save folder")

elif sys.argv[1]=="xlsx":
    def perturb_string(string):
        perturbed_chars = []
        for char in string:
            if char.isalpha():
                perturbed_chars.append(chr(ord(char) + random.randint(1, 5)))
            elif char.isdigit():
                perturbed_chars.append(str((int(char) + random.randint(1, 5)) % 10))
            else:
                perturbed_chars.append(char)
        return ''.join(perturbed_chars)


    def perturb_names(sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    words = cell.value.split()
                    perturbed_words = [perturb_string(word) for word in words]
                    perturbed_value = ' '.join(perturbed_words)
                    cell.value = perturbed_value


    def perturb_numbers(sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.value = perturb_string(str(cell.value))


    def perturb_phone_numbers(sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(char.isdigit() for char in cell.value):
                    perturbed_value = ''.join(perturb_string(char) if char.isdigit() else char for char in cell.value)
                    cell.value = perturbed_value


    def perturb_email_addresses(sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and '@' in cell.value:
                    parts = cell.value.split('@')
                    perturbed_username = perturb_string(parts[0])
                    perturbed_value = perturbed_username + '@' + parts[1]
                    cell.value = perturbed_value


    def perturb_xlsx_file(filename):
        workbook = load_workbook(filename)
        for sheet in workbook:
            perturb_names(sheet)
            perturb_numbers(sheet)
            perturb_phone_numbers(sheet)
            perturb_email_addresses(sheet)
        workbook.save(sys.argv[2])


    # Usage example
    filename = sys.argv[2]
    perturb_xlsx_file(filename)
    print("Done")